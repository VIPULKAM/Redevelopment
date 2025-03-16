#!/usr/bin/env python3
"""
Excel Generator for Redevelopment Profit Calculator

This script generates a Microsoft Excel file with all the formulas and calculations
for the redevelopment profit calculator. The Excel file includes:
- Input section for parameters
- Area calculations
- Cost calculations
- Revenue and profit calculations
- Salable flats and surplus corpus calculations
- Summary dashboard
- Documentation
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, PieChart

def create_redevelopment_excel():
    """Create an Excel file with redevelopment profit calculations."""
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Create the sheets
    inputs_sheet = wb.create_sheet("Inputs")
    area_sheet = wb.create_sheet("Area Calculations")
    cost_sheet = wb.create_sheet("Cost Calculations")
    profit_sheet = wb.create_sheet("Profit & Surplus")
    summary_sheet = wb.create_sheet("Summary Dashboard")
    docs_sheet = wb.create_sheet("Documentation")
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    subheader_font = Font(name='Calibri', size=11, bold=True)
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    input_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    formula_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    result_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Center alignment
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # -------------------------------------------------------------
    # INPUTS SHEET
    # -------------------------------------------------------------
    inputs_sheet.title = "Inputs"
    
    # Add title
    inputs_sheet.merge_cells('A1:D1')
    inputs_sheet['A1'] = "REDEVELOPMENT PROFIT CALCULATOR - INPUTS"
    inputs_sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
    inputs_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers
    headers = ['Parameter', 'Value', 'Unit', 'Description']
    for col, header in enumerate(headers):
        cell = inputs_sheet.cell(row=3, column=col+1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Input parameters
    parameters = [
        # Land & Member Parameters
        ["Land & Member Parameters", "", "", ""],
        ["Land Area", 10, "Guntha", "Total land area of the society"],
        ["Total Members", 40, "flats", "Number of existing flats/members"],
        ["Carpet Area per Member", 500, "sqft", "Current carpet area per flat"],
        ["Extra Carpet Percentage", 30, "%", "Additional area to be offered"],
        
        # Development Parameters
        ["Development Parameters", "", "", ""],
        ["FSI", 2.5, "", "Floor Space Index (Chatai Kshetra Nirdeshank)"],
        ["Ancillary Percentage", 60, "%", "Percentage of ancillary area allowed on top of FSI"],
        ["Green Building Bonus", 7, "%", "Percentage of additional FSI for green building"],
        ["Self-Redevelopment Bonus", 10, "%", "Percentage of additional FSI for self-redevelopment"],
        
        # Cost Parameters
        ["Cost Parameters", "", "", ""],
        ["Monthly Rent per Flat", 15000, "₹", "Monthly rent paid during construction"],
        ["Rent Duration", 36, "months", "Duration of rent payment"],
        ["Relocation Cost per Member", 20000, "₹", "Relocation and brokerage cost per member"],
        ["Construction Cost", 3000, "₹/sqft", "Construction cost per square foot"],
        ["TMC Premium", 50000000, "₹", "Premium to be paid to municipal corporation"],
        ["Bank Interest", 50000000, "₹", "Bank interest amount"],
        
        # Revenue Parameters
        ["Revenue Parameters", "", "", ""],
        ["Market Rate", 17500, "₹/sqft", "Market rate per square foot for selling"],
        ["Average New Flat Size", 750, "sqft", "Average size of new salable flats"],
        
        # Project Type
        ["Project Type", "", "", ""],
        ["Is Self-Redevelopment", "Yes", "", "Self-redevelopment or builder project"],
        ["Profit Sharing with Developer", 50, "%", "Only applicable if not self-redevelopment"]
    ]
    
    row = 4
    for param in parameters:
        # Check if this is a section header
        if param[1] == "" and param[2] == "" and param[3] == "":
            # Section header
            inputs_sheet.merge_cells(f'A{row}:D{row}')
            cell = inputs_sheet.cell(row=row, column=1)
            cell.value = param[0]
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            for col in range(1, 5):
                inputs_sheet.cell(row=row, column=col).border = thin_border
        else:
            # Parameter row
            for col, value in enumerate(param):
                cell = inputs_sheet.cell(row=row, column=col+1)
                cell.value = value
                if col == 0:  # Parameter name
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col == 1:  # Value
                    cell.alignment = right_align
                    cell.fill = input_fill
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
        row += 1
    
    # Named ranges for inputs
    input_ranges = {
        "land_area": "B5",
        "total_members": "B6",
        "carpet_area_per_member": "B7",
        "extra_carpet_percentage": "B8",
        "fsi": "B11",
        "ancillary_percentage": "B12",
        "green_building_bonus": "B13",
        "self_redevelopment_bonus": "B14",
        "monthly_rent": "B17",
        "rent_duration": "B18",
        "relocation_cost": "B19",
        "construction_cost": "B20",
        "tmc_premium": "B21",
        "bank_interest": "B22",
        "market_rate": "B25",
        "avg_flat_size": "B26",
        "is_self_redevelopment": "B29",
        "profit_sharing": "B30"
    }
    
    # Create named ranges
    for name, cell_ref in input_ranges.items():
        wb.create_named_range(name, inputs_sheet, cell_ref)
    
    # Add data validation for Yes/No field
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv.add(inputs_sheet["B29"])
    inputs_sheet.add_data_validation(dv)
    
    # Adjust column widths
    inputs_sheet.column_dimensions['A'].width = 30
    inputs_sheet.column_dimensions['B'].width = 15
    inputs_sheet.column_dimensions['C'].width = 10
    inputs_sheet.column_dimensions['D'].width = 50
    
    # -------------------------------------------------------------
    # AREA CALCULATIONS SHEET
    # -------------------------------------------------------------
    area_sheet.title = "Area Calculations"
    
    # Add title
    area_sheet.merge_cells('A1:E1')
    area_sheet['A1'] = "AREA CALCULATIONS"
    area_sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
    area_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers
    area_headers = ['Step', 'Calculation', 'Formula', 'Value', 'Unit']
    for col, header in enumerate(area_headers):
        cell = area_sheet.cell(row=3, column=col+1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Area calculation steps
    area_calculations = [
        # Basic Land Area
        ["1", "Land Area in Square Meters", "=land_area*101.17", "", "sqm"],
        
        # Carpet Areas
        ["2", "Current Total Carpet Area", "=carpet_area_per_member*total_members", "", "sqft"],
        ["2", "Offered Carpet Area (with extra)", "=C5*(1+extra_carpet_percentage/100)", "", "sqft"],
        
        # FSI Area
        ["3", "Basic FSI Area", "=C4*fsi", "", "sqm"],
        ["3", "Ancillary Area", "=C7*(ancillary_percentage/100)", "", "sqm"],
        ["3", "Total Buildable Area (sqm)", "=C7+C8", "", "sqm"],
        ["3", "Total Buildable Area (sqft)", "=C9*10.764", "", "sqft"],
        
        # Green Building Bonus
        ["4", "Green Building Bonus Area", "=C10*(green_building_bonus/100)", "", "sqft"],
        ["4", "Area with Green Bonus", "=C10+C11", "", "sqft"],
        
        # Self-Redevelopment Bonus
        ["5", "Self-Redevelopment Bonus", "=IF(is_self_redevelopment=\"Yes\",C12*(self_redevelopment_bonus/100),0)", "", "sqft"],
        ["5", "Total Final Area", "=C12+C13", "", "sqft"],
        
        # Sellable Area
        ["6", "Builder Sellable Area", "=C12-C6", "", "sqft"]
    ]
    
    row = 4
    for calc in area_calculations:
        for col, value in enumerate(calc):
            cell = area_sheet.cell(row=row, column=col+1)
            if col == 2:  # Formula column
                cell.value = value
                cell.fill = formula_fill
            elif col == 3:  # Value column
                if row == 4:
                    cell.value = "=ROUND(" + value + ",2)"
                else:
                    cell.value = "=ROUND(" + value + ",2)"
                cell.fill = result_fill
            else:
                cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        row += 1
    
    # Adjust column widths
    area_sheet.column_dimensions['A'].width = 5
    area_sheet.column_dimensions['B'].width = 30
    area_sheet.column_dimensions['C'].width = 40
    area_sheet.column_dimensions['D'].width = 15
    area_sheet.column_dimensions['E'].width = 10
    
    # Create Bar Chart for Areas
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Area Distribution (sqft)"
    chart1.y_axis.title = "Square Feet"
    
    data = Reference(area_sheet, min_col=4, min_row=4, max_row=15)
    cats = Reference(area_sheet, min_col=2, min_row=4, max_row=15)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.shape = 4
    area_sheet.add_chart(chart1, "G4")
    
    # -------------------------------------------------------------
    # COST CALCULATIONS SHEET
    # -------------------------------------------------------------
    cost_sheet.title = "Cost Calculations"
    
    # Add title
    cost_sheet.merge_cells('A1:E1')
    cost_sheet['A1'] = "COST CALCULATIONS"
    cost_sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
    cost_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers
    cost_headers = ['Step', 'Calculation', 'Formula', 'Value', 'Unit']
    for col, header in enumerate(cost_headers):
        cell = cost_sheet.cell(row=3, column=col+1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Cost calculation steps
    cost_calculations = [
        # Accommodation Costs
        ["7", "Rent Cost", "=monthly_rent*total_members*rent_duration", "", "₹"],
        ["7", "Relocation Cost", "=relocation_cost*total_members", "", "₹"],
        ["7", "Total Accommodation Cost", "=C4+C5", "", "₹"],
        
        # Construction Costs
        ["8", "Construction Cost", "='Area Calculations'!C12*construction_cost", "", "₹"],
        ["8", "Total Construction Cost", "=C7+C6", "", "₹"],
        
        # Total Project Cost
        ["9", "TMC Premium", "=tmc_premium", "", "₹"],
        ["9", "Bank Interest", "=bank_interest", "", "₹"],
        ["9", "Total Project Cost", "=C8+C9+C10", "", "₹"]
    ]
    
    row = 4
    for calc in cost_calculations:
        for col, value in enumerate(calc):
            cell = cost_sheet.cell(row=row, column=col+1)
            if col == 2:  # Formula column
                cell.value = value
                cell.fill = formula_fill
            elif col == 3:  # Value column
                if value:
                    cell.value = "=ROUND(" + value + ",2)"
                cell.fill = result_fill
                cell.number_format = '#,##0'
            else:
                cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        row += 1
    
    # Adjust column widths
    cost_sheet.column_dimensions['A'].width = 5
    cost_sheet.column_dimensions['B'].width = 30
    cost_sheet.column_dimensions['C'].width = 40
    cost_sheet.column_dimensions['D'].width = 15
    cost_sheet.column_dimensions['E'].width = 10
    
    # Create Pie Chart for Costs
    chart2 = PieChart()
    chart2.title = "Cost Distribution"
    
    labels = Reference(cost_sheet, min_col=2, min_row=4, max_row=10)
    data = Reference(cost_sheet, min_col=4, min_row=4, max_row=10)
    chart2.add_data(data, titles_from_data=False)
    chart2.set_categories(labels)
    chart2.dataLabels = openpyxl.chart.label.DataLabelList()
    chart2.dataLabels.showPercent = True
    
    cost_sheet.add_chart(chart2, "G4")
    
    # -------------------------------------------------------------
    # PROFIT & SURPLUS SHEET
    # -------------------------------------------------------------
    profit_sheet.title = "Profit & Surplus"
    
    # Add title
    profit_sheet.merge_cells('A1:E1')
    profit_sheet['A1'] = "PROFIT & SURPLUS CALCULATIONS"
    profit_sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
    profit_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Column headers
    profit_headers = ['Step', 'Calculation', 'Formula', 'Value', 'Unit']
    for col, header in enumerate(profit_headers):
        cell = profit_sheet.cell(row=3, column=col+1)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Profit & Surplus calculation steps
    profit_calculations = [
        # Project Value
        ["10", "Project Value", "='Area Calculations'!C14*market_rate", "", "₹"],
        
        # Profit
        ["11", "Total Profit", "=C4-'Cost Calculations'!C11", "", "₹"],
        
        # Salable Flats
        ["12", "Number of Salable Flats", "='Area Calculations'!C15/avg_flat_size", "", "flats"],
        
        # Profit Distribution
        ["13", "Developer's Profit", "=IF(is_self_redevelopment=\"No\",C5*(profit_sharing/100),0)", "", "₹"],
        ["13", "Society's Profit", "=C5-C8", "", "₹"],
        
        # Per Member Profit
        ["14", "Profit per Member", "=C9/total_members", "", "₹"],
        
        # Surplus Corpus
        ["15", "Surplus Corpus for Existing Members", "=C9", "", "₹"],
        
        # ROI
        ["16", "Return on Investment (ROI)", "=(C5/'Cost Calculations'!C11)*100", "", "%"]
    ]
    
    row = 4
    for calc in profit_calculations:
        for col, value in enumerate(calc):
            cell = profit_sheet.cell(row=row, column=col+1)
            if col == 2:  # Formula column
                cell.value = value
                cell.fill = formula_fill
            elif col == 3:  # Value column
                if value:
                    if "%" in calc[4]:  # If percentage
                        cell.value = "=ROUND(" + value + ",2)"
                        cell.number_format = '0.00"%"'
                    else:
                        cell.value = "=ROUND(" + value + ",2)"
                        cell.number_format = '#,##0'
                cell.fill = result_fill
            else:
                cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        row += 1
    
    # Add Crore Conversion Section
    profit_sheet['A13'] = "Conversion to Crores"
    profit_sheet['A13'].font = subheader_font
    profit_sheet['A13'].fill = subheader_fill
    profit_sheet.merge_cells('A13:E13')
    
    crore_conversions = [
        ["", "Total Profit in Crores", "=C5/10000000", "", "Cr"],
        ["", "Society's Profit in Crores", "=C9/10000000", "", "Cr"],
        ["", "Profit per Member in Crores", "=C10/10000000", "", "Cr"]
    ]
    
    row = 14
    for calc in crore_conversions:
        for col, value in enumerate(calc):
            cell = profit_sheet.cell(row=row, column=col+1)
            if col == 2:  # Formula column
                cell.value = value
                cell.fill = formula_fill
            elif col == 3:  # Value column
                if value:
                    cell.value = "=ROUND(" + value + ",2)"
                cell.fill = result_fill
                cell.number_format = '#,##0.00'
            else:
                cell.value = value
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        row += 1
    
    # Adjust column widths
    profit_sheet.column_dimensions['A'].width = 5
    profit_sheet.column_dimensions['B'].width = 40
    profit_sheet.column_dimensions['C'].width = 40
    profit_sheet.column_dimensions['D'].width = 15
    profit_sheet.column_dimensions['E'].width = 10
    
    # -------------------------------------------------------------
    # SUMMARY DASHBOARD SHEET
    # -------------------------------------------------------------
    summary_sheet.title = "Summary Dashboard"
    
    # Add title
    summary_sheet.merge_cells('A1:D1')
    summary_sheet['A1'] = "REDEVELOPMENT PROFIT SUMMARY"
    summary_sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
    summary_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Project Type Section
    summary_sheet['A3'] = "PROJECT TYPE:"
    summary_sheet['A3'].font = subheader_font
    
    summary_sheet['B3'] = "=IF(is_self_redevelopment=\"Yes\",\"Self-Redevelopment\",\"Builder Redevelopment\")"
    summary_sheet['B3'].font = Font(size=12, bold=True, color="0000FF")
    
    # Key Metrics Section
    summary_sheet['A5'] = "KEY METRICS"
    summary_sheet['A5'].font = subheader_font
    summary_sheet['A5'].fill = subheader_fill
    summary_sheet.merge_cells('A5:D5')
    
    metrics = [
        ["Land Area", "=land_area", "Guntha"],
        ["Total Members", "=total_members", "flats"],
        ["Total Final Area", "='Area Calculations'!D14", "sqft"],
        ["Builder Sellable Area", "='Area Calculations'!D15", "sqft"],
        ["Total Project Cost", "='Cost Calculations'!D11", "₹"],
        ["Project Value", "='Profit & Surplus'!D4", "₹"],
        ["Total Profit", "='Profit & Surplus'!D5", "₹"],
        ["Number of Salable Flats", "=ROUND('Profit & Surplus'!D7,0)", "flats"],
        ["Return on Investment (ROI)", "='Profit & Surplus'!D12", "%"]
    ]
    
    row = 6
    for metric in metrics:
        summary_sheet.cell(row=row, column=1).value = metric[0]
        summary_sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        summary_sheet.cell(row=row, column=2).value = metric[1]
        summary_sheet.cell(row=row, column=2).number_format = '#,##0.00'
        summary_sheet.cell(row=row, column=2).alignment = right_align
        
        summary_sheet.cell(row=row, column=3).value = metric[2]
        summary_sheet.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # Member Benefits Section
    summary_sheet['A16'] = "MEMBER BENEFITS"
    summary_sheet['A16'].font = subheader_font
    summary_sheet['A16'].fill = subheader_fill
    summary_sheet.merge_cells('A16:D16')
    
    benefits = [
        ["Current Carpet Area", "=carpet_area_per_member", "sqft/member"],
        ["Offered Carpet Area", "='Area Calculations'!D6/total_members", "sqft/member"],
        ["Area Increase", "=('Area Calculations'!D6-'Area Calculations'!D5)/total_members", "sqft/member"],
        ["Area Increase Percentage", "=('Area Calculations'!D6/'Area Calculations'!D5-1)*100", "%"],
        ["Surplus per Member", "='Profit & Surplus'!D10", "₹"],
        ["Surplus per Member (Crores)", "='Profit & Surplus'!D16", "Cr"]
    ]
    
    row = 17
    for benefit in benefits:
        summary_sheet.cell(row=row, column=1).value = benefit[0]
        summary_sheet.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        summary_sheet.cell(row=row, column=2).value = benefit[1]
        if "%" in benefit[2]:
            summary_sheet.cell(row=row, column=2).number_format = '0.00"%"'
        else:
            summary_sheet.cell(row=row, column=2).number_format = '#,##0.00'
        summary_sheet.cell(row=row, column=2).alignment = right_align
        
        summary_sheet.cell(row=row, column=3).value = benefit[2]
        summary_sheet.cell(row=row, column=3).alignment = Alignment(horizontal='left', vertical='center')
        
        row += 1
    
    # Add notice about builder profit
    if_builder = "=IF(is_self_redevelopment=\"No\","
    builder_profit_cell = if_builder + "\"Builder's Profit per Flat: \" & TEXT('Profit & Surplus'!D8/total_members/10000000,\"0.00\") & \" Cr\",\"\")"
    
    summary_sheet['A24'] = builder_profit_cell
    summary_sheet['A24'].font = Font(size=12, bold=True, color="FF0000")
    summary_sheet.merge_cells('A24:D24')
    
    # Add warning about self-redevelopment
    summary_sheet['A26'] = "NOTE: Self-redevelopment can yield significantly higher returns but requires professional management and commitment from society members."
    summary_sheet['A26'].font = Font(italic=True)
    summary_sheet.merge_cells('A26:D26')
    
    # Adjust column widths
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 15
    summary_sheet.column_dimensions['D'].width = 30
    
    # -------------------------------------------------------------
    # DOCUMENTATION SHEET
    # -------------------------------------------------------------
    docs_sheet.title = "Documentation"
    
    # Add title
    docs_sheet.merge_cells('A1:C1')
    docs_sheet['A1'] = "DOCUMENTATION & HELP"
    docs_sheet['A1'].font = Font(name='Calibri', size=16, bold=True)
    docs_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add sections
    sections = [
        ["OVERVIEW", "", ""],
        ["This spreadsheet calculates the financial aspects of a housing society redevelopment project. It shows the profit potential, distribution of benefits, and number of salable flats that can be generated.", "", ""],
        ["", "", ""],
        ["KEY CONCEPTS", "", ""],
        ["Guntha", "A unit of land measurement used in parts of India. 1 Guntha = 101.17 square meters.", ""],
        ["FSI (Floor Space Index)", "The ratio of a building's total floor area to the size of the land upon which it is built.", "Also known as FAR (Floor Area Ratio)"],
        ["Ancillary Area", "Additional construction area allowed beyond the basic FSI, typically 60% of the FSI area.", ""],
        ["Green Building Bonus", "Additional FSI incentive (typically 7%) for implementing green building features.", ""],
        ["Self-Redevelopment Bonus", "Additional FSI incentive (typically 10%) for societies undertaking self-redevelopment.", ""],
        ["", "", ""],
        ["HOW TO USE THIS CALCULATOR", "", ""],
        ["1. Enter all parameters in the 'Inputs' sheet.", "", ""],
        ["2. Review the detailed calculations in the 'Area Calculations', 'Cost Calculations', and 'Profit & Surplus' sheets.", "", ""],
        ["3. See the summary of results in the 'Summary Dashboard' sheet.", "", ""],
        ["", "", ""],
        ["IMPORTANT NOTES", "", ""],
        ["- All calculations are based on standard redevelopment practices in Maharashtra, India.", "", ""],
        ["- The actual profit may vary based on market conditions, approvals, and project execution.", "", ""],
        ["- Self-redevelopment requires professional project management and commitment from society members.", "", ""],
        ["- This calculator reveals the true profit potential of redevelopment projects, which is often not transparent in builder offers.", "", ""],
        ["", "", ""],
        ["FORMULA EXPLANATIONS", "", ""],
        ["Area Conversion", "1 Guntha = 101.17 square meters; 1 square meter = 10.764 square feet", ""],
        ["Offered Carpet Area", "Current carpet area × (1 + extra percentage / 100)", ""],
        ["Builder Sellable Area", "Total area with green bonus - Offered carpet area", ""],
        ["Number of Salable Flats", "Builder sellable area ÷ Average new flat size", ""],
        ["Profit Distribution", "In builder redevelopment: Builder gets profit sharing %, society gets the rest", ""],
        ["Surplus Corpus", "Society's profit that will be distributed among existing members", ""],
        ["", "", ""],
        ["CREATED BY", "", ""],
        ["This calculator was created to bring transparency to redevelopment economics and reveal the true profit potential for housing societies.", "", ""],
        ["For educational purposes only. Please consult professionals before making redevelopment decisions.", "", ""]
    ]
    
    row = 3
    for section in sections:
        for col, value in enumerate(section):
            if value:
                docs_sheet.cell(row=row, column=col+1).value = value
                if row in [3, 6]:  # Section headers
                    docs_sheet.cell(row=row, column=col+1).font = subheader_font
                    docs_sheet.cell(row=row, column=col+1).fill = subheader_fill
                elif col == 0 and len(value) < 50:  # Item headers
                    docs_sheet.cell(row=row, column=col+1).font = Font(bold=True)
        row += 1
    
    # Adjust column widths
    docs_sheet.column_dimensions['A'].width = 30
    docs_sheet.column_dimensions['B'].width = 60
    docs_sheet.column_dimensions['C'].width = 30

    # Save the workbook
    filename = "Redevelopment_Profit_Calculator.xlsx"
    wb.save(filename)
    print(f"Excel file '{filename}' created successfully!")
    return filename


def main():
    """Main function to run the Excel generator."""
    try:
        print("REDEVELOPMENT PROFIT CALCULATOR - EXCEL GENERATOR")
        print("=" * 50)
        print("This script will generate an Excel file with all calculations.")
        print("The Excel file will have multiple sheets for inputs, calculations, and summary.")

        # Create the Excel file
        filename = create_redevelopment_excel()

        print("\nEXCEL FILE CREATED SUCCESSFULLY!")
        print(f"Filename: {filename}")
        print("\nThe Excel file contains the following sheets:")
        print("1. Inputs - Enter all parameters here")
        print("2. Area Calculations - Detailed area calculations")
        print("3. Cost Calculations - All cost components")
        print("4. Profit & Surplus - Profit calculations and surplus distribution")
        print("5. Summary Dashboard - Key metrics and results")
        print("6. Documentation - Help and explanations")

        print("\nInstructions:")
        print("1. Open the Excel file in Microsoft Excel")
        print("2. Start by entering your parameters in the 'Inputs' sheet")
        print("3. All calculations will update automatically")
        print("4. Review the results in the 'Summary Dashboard'")

    except Exception as e:
        print(f"\nAn error occurred: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
